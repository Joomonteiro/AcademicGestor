from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Mysheet") # insert at the end (default)
ws2 = wb.create_sheet("Mysheet", 0) # insert at first position
ws3 = wb.create_sheet("Mysheet", -1) # insert at the penultimate position
ws.title = "New Title"
ws.sheet_properties.tabColor = "1072BA"
ws3 = wb["New Title"]
print(ws3)
print(wb.sheetnames)
for sheet in wb:   
    print(sheet.title)
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)
        cell.value = 'hello, world'
        print(cell.value)

wb.save('balances.xlsx')
